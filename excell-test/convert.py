import os
import sys
import csv
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl


def excel_to_csv(ws, path, delimiter=","):
    ext = ".csv" if delimiter == "," else ".txt"
    out = Path(path).stem + ext
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, delimiter=delimiter)
        for row in ws.iter_rows(values_only=True):
            w.writerow([str(c) if c is not None else "" for c in row])
    return out


def main():
    path = input("Caminho do ficheiro Excel: ").strip().strip('"').strip("'")
    if not os.path.isfile(path):
        print("Ficheiro não encontrado.")
        return

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    print(f"\nFolha activa: {ws.title}  |  Linhas: {ws.max_row}  |  Colunas: {ws.max_column}\n")

    choice = input("Converter para [C]SV ou [T]XT? ").strip().lower()
    if choice == "c":
        out = excel_to_csv(ws, path, ",")
        print(f"CSV salvo: {out}")
    elif choice == "t":
        out = excel_to_csv(ws, path, ";")
        print(f"TXT salvo: {out}")
    else:
        print("Opção inválida.")

    wb.close()


if __name__ == "__main__":
    main()
